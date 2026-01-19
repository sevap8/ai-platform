"""
LangChain-compatible file loader for the AI Platform.

This module provides a file loader implementation compatible with LangChain
that can handle various file types including text, Excel, and PDF files.
"""

from typing import List, Optional, Union
from pathlib import Path
import pandas as pd
from io import StringIO, BytesIO
import unicodedata
from openpyxl import load_workbook
import orjson
import yaml
from defusedxml import ElementTree
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.utils import open_filename
import xlrd
from langchain_core.documents import Document
from langchain_community.document_loaders.base import BaseLoader


class FileLoader(BaseLoader):
    """LangChain-compatible loader for processing various file types including text, Excel, and PDF files."""

    def __init__(
        self,
        file_path: Union[str, Path],
        file_uuid: Optional[str] = None,
        password: str = '',
        maxpages: int = 0,
        caching: bool = True,
        codec: str = 'utf-8',
        laparams: Optional[LAParams] = None
    ):
        """
        Initialize the FileLoader.

        Args:
            file_path: Path to the file to be loaded
            file_uuid: UUID of the file (optional)
            password: Password for encrypted PDF files
            maxpages: Maximum number of pages to process in PDF files
            caching: Whether to use caching for PDF processing
            codec: Text encoding to use
            laparams: Layout parameters for PDF text extraction
        """
        self.file_path = Path(file_path)
        self.file_uuid = file_uuid or str(file_path)
        self.password = password
        self.maxpages = maxpages
        self.caching = caching
        self.codec = codec
        self.laparams = laparams or LAParams()

    def extract_excel_to_md(self, file_path: str) -> str:
        """
        Extract content from Excel file and return it in Markdown format.
        
        Args:
            file_path: Path to the Excel file (.xlsx or .xls)
        
        Returns:
            str: Content of the file in Markdown format
        """
        file_name = self.file_path.name
        all_md_content = ""
        file_extension = self.file_path.suffix.lower()

        if file_extension == ".xlsx":
            wb = load_workbook(filename=file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                data = sheet.values
                try:
                    cols = next(data)
                except StopIteration:
                    continue
                df = pd.DataFrame(data, columns=cols)
                df.dropna(how="all", inplace=True)

                md_content = f"## Sheet: {sheet_name}\n\n"
                md_content += "| " + " | ".join(str(col) for col in cols) + " |\n"
                md_content += "| " + " | ".join(["---"] * len(cols)) + " |\n"
                
                for index, row in df.iterrows():
                    md_content += (
                        "| "
                        + " | ".join(
                            [str(cell) if pd.notna(cell) else "" for cell in row]
                        )
                        + " |\n"
                    )
                all_md_content += md_content + "\n"

        elif file_extension == ".xls":
            excel_file = pd.ExcelFile(file_path, engine="xlrd")
            for sheet_name in excel_file.sheet_names:
                df = excel_file.parse(sheet_name=sheet_name)
                df.dropna(how="all", inplace=True)

                md_content = f"## {sheet_name}\n\n"
                md_content += "| " + " | ".join(df.columns) + " |\n"
                md_content += "| " + " | ".join(["---"] * len(df.columns)) + " |\n"
                
                for _, row in df.iterrows():
                    md_content += (
                        "| "
                        + " | ".join([str(cell) if pd.notna(cell) else "" for cell in row])
                        + " |\n"
                    )

                all_md_content += md_content + "\n"
        else:
            raise ValueError(f"Unsupported file extension: {file_extension}")

        return all_md_content

    def parse_pdf_to_text(self, file_path: str) -> List[str]:
        """
        Parse PDF file and extract text.
        
        Args:
            file_path: Path to the PDF file.
        
        Returns:
            List of strings, each representing text from a page.
        """
        texts = []
        with open_filename(file_path, "rb") as fp:
            rsrcmgr = PDFResourceManager(caching=self.caching)
            for idx, page in enumerate(PDFPage.get_pages(
                    fp,
                    maxpages=self.maxpages,
                    password=self.password,
                    caching=self.caching,
            )):
                with StringIO() as output_string:
                    device = TextConverter(rsrcmgr, output_string, codec=self.codec,
                                           laparams=self.laparams)
                    interpreter = PDFPageInterpreter(rsrcmgr, device)
                    interpreter.process_page(page)
                    page_text = output_string.getvalue()
                    normalized_text = unicodedata.normalize("NFKD", page_text).replace('\n\n','\n')
                    texts.append(normalized_text)
        
        return texts

    def load(self) -> List[Document]:
        """
        Load and process the file based on its type.
        
        Returns:
            List of LangChain Document objects
        """
        file_ext = self.file_path.suffix.lower()
        
        if file_ext == ".pdf":
            texts = self.parse_pdf_to_text(str(self.file_path))
            docs = []
            for i, text in enumerate(texts):
                docs.append(Document(
                    page_content=text,
                    metadata={
                        "source": str(self.file_path),
                        "page": i,
                        "file_uuid": self.file_uuid
                    }
                ))
            return docs
        elif file_ext in [".xlsx", ".xls"]:
            content = self.extract_excel_to_md(str(self.file_path))
            return [Document(
                page_content=content,
                metadata={
                    "source": str(self.file_path),
                    "file_uuid": self.file_uuid
                }
            )]
        elif file_ext in ['.txt', '.py', '.js', '.ts', '.jsx', '.tsx', '.csv', '.json', '.yaml', '.yml', '.xml']:
            with open(self.file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Handle structured formats
            if file_ext == '.json':
                parsed_data = orjson.loads(content)
                content = orjson.dumps(parsed_data).decode("utf-8")
            elif file_ext in ['.yaml', '.yml']:
                parsed_data = yaml.safe_load(content)
                content = str(parsed_data)
            elif file_ext == '.xml':
                xml_element = ElementTree.fromstring(content)
                content = ElementTree.tostring(xml_element, encoding="unicode")
            
            return [Document(
                page_content=content,
                metadata={
                    "source": str(self.file_path),
                    "file_uuid": self.file_uuid
                }
            )]
        else:
            raise ValueError(f"Unsupported file extension: {file_ext}")
